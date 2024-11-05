import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.http import HttpResponse
from .models import Preliminar, Asistencia, Descargue  # Asegúrate de importar los modelos correctos
from django.shortcuts import get_object_or_404
from .models import Campania  # Asegúrate de importar el modelo Campania
from django.utils import timezone  # Importa timezone para obtener la fecha actual

def export_preliminar_excel(request, pk):
    # Obtener la campaña específica
    campania = get_object_or_404(Campania, id=pk)

    # Crea un nuevo workbook y selecciona la hoja activa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Detallado Preliminar"

    # Agregar el título
    title = f"Planilla {campania.nombre}"
    ws.append([title])  # Agrega el título en una nueva fila
    title_font = Font(bold=True, size=14)  # Estilo para el título
    ws["A1"].font = title_font  # Aplica el estilo a la celda A1
    ws["A1"].alignment = Alignment(horizontal='center', vertical='center')  # Centra el texto
    ws.merge_cells('A1:L1')  # Combina las celdas de A1 a L1 para el título

    # Agregar fecha y tonelaje con otros valores
    date_str = timezone.now().strftime("%Y-%m-%d")  # Obtiene la fecha actual en formato YYYY-MM-DD
    ws.append([f"Fecha: {date_str}"])  # Agrega la fecha en una nueva fila

    # Agrega la línea con tonelaje, vacaciones, aguinaldo, séptimo y cambio
    ws.append([
        f"Tonelaje: {campania.tonelaje}", 
        f"Vacaciones: {campania.vacaciones}", 
        f"Aguinaldo: {campania.aguinaldo}", 
        f"Séptimo: {campania.septimo}", 
        f"Cambio: {campania.cambio}"
    ])

    # Estilo para la fecha y los nuevos valores
    for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=1):  # Ajusta para las filas de fecha y valores
        for cell in row:
            cell.font = Font(italic=True)  # Estilo en cursiva para fecha y tonelaje
            cell.alignment = Alignment(horizontal='left', vertical='top')  # Cambié a 'top'

    # Agrega el encabezado en la fila 4
    headers = [
        "Campaña", "Cargo", "Persona", "Descargue", "Días trabajado",
        "Salario Base", "Vacaciones", "Aguinaldo", "Indemnización",
        "INSS Laboral", "INSS Patronal", "Total salario"
    ]
    ws.append(headers)

    # Estilo para el encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[4]:  # Cambia de 2 a 4 para aplicar el estilo en la fila de encabezados
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Establecer el ancho de las columnas
    column_widths = [15, 15, 20, 10, 5, 15, 15, 15, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Obtiene los datos filtrando por la campaña específica
    preliminares = Preliminar.objects.filter(campania=campania)  # Filtrar por campaña

    # Recorre cada registro de Preliminar para añadir una fila por persona
    for preliminar in preliminares:
        # Obtener asistencia y descargue para la campaña y el personal específicos
        asistencia = Asistencia.objects.filter(campania=preliminar.campania, personal=preliminar.personal)
        descargue = Descargue.objects.filter(campania=preliminar.campania)

        # Inicialización de variables para cálculos
        sumatoria_monto = 0
        sumatoria_dias = 0
        sumatoria_salario_dia = 0
        sumatoria_salario_base = 0
        sumatoria_vacaciones, sumatoria_aguinaldo, sumatoria_indemnizacion = 0, 0, 0  # Asignación corregida


        for a in asistencia:
            for d in descargue:
                if a.fecha == d.fecha:
                    # Cuenta de personal con mismo cargo en esa fecha
                    personal_count = Asistencia.objects.filter(
                        campania=preliminar.campania,
                        fecha=a.fecha,
                        cargo=preliminar.cargo
                    ).count()

                    # Cálculo de valores por día y sumas totales
                    salario_por_dia = (preliminar.total / personal_count) * d.monto
                    salario_base = (preliminar.sb() / personal_count) * d.monto

                    # Ajustar los métodos para asegurar que devuelvan valores numéricos
                    try:
                        vacaciones = (preliminar.vacaciones()['monto'] / personal_count) * d.monto
                    except (KeyError, TypeError):
                        vacaciones = 0  # Valor por defecto si no se puede obtener

                    try:
                        aguinaldo = (preliminar.aguinaldo()['monto'] / personal_count) * d.monto
                    except (KeyError, TypeError):
                        aguinaldo = 0  # Valor por defecto si no se puede obtener

                    try:
                        indemnizacion = (preliminar.indemnizacion()['monto'] / personal_count) * d.monto
                    except (KeyError, TypeError):
                        indemnizacion = 0  # Valor por defecto si no se puede obtener

                    # Agregar a sumatorias
                    sumatoria_salario_dia += salario_por_dia
                    sumatoria_salario_base += salario_base
                    sumatoria_vacaciones += vacaciones
                    sumatoria_aguinaldo += aguinaldo
                    sumatoria_indemnizacion += indemnizacion
                    sumatoria_monto += d.monto
                    sumatoria_dias += 1

        # Multiplica los totales por campania.cambio
        total_dias_trabajados = sumatoria_dias
        total_salario_base = round(sumatoria_salario_base * campania.cambio, 2)
        total_vacaciones = round(sumatoria_vacaciones * campania.cambio, 2)
        total_aguinaldo = round(sumatoria_aguinaldo * campania.cambio, 2)
        total_indemnizacion = round(sumatoria_indemnizacion * campania.cambio, 2)
        total_inss_laboral = round((total_salario_base + total_vacaciones) * 0.07, 2)
        total_inss_patronal = round(total_salario_base * 0.125, 2)
        total_salario = round(total_salario_base + total_vacaciones + total_aguinaldo + total_indemnizacion, 2)

        # Agregar fila al Excel con los datos calculados
        row = [
            preliminar.campania.nombre,               # Campaña
            preliminar.cargo.nombre,                  # Cargo
            preliminar.personal.nombre,                # Persona
            round(sumatoria_monto, 2),                # Monto total calculado
            total_dias_trabajados,                    # Días total
            total_salario_base,                       # Salario Base totals
            total_vacaciones,                         # Vacaciones total
            total_aguinaldo,                          # Aguinaldo total
            total_indemnizacion,                      # Indemnización total
            total_inss_laboral,                       # INSS Laboral
            total_inss_patronal,                      # INSS Patronal
            total_salario                               # Total Salario
        ]
        ws.append(row)

    # Estilo para las filas de datos
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):  # Cambia a 5 para las filas de datos
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Configura la respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename=report_preliminar_{pk}.xlsx'
    wb.save(response)  # Guarda el workbook en la respuesta

    return response
